import customtkinter as ctk
from datetime import date, timedelta
from typing import Callable


class DateRangePicker(ctk.CTkFrame):
    def __init__(self, parent, callback: Callable = None, **kwargs):
        super().__init__(parent, fg_color="transparent", **kwargs)
        self.callback = callback
        self.start_date = date.today().replace(day=1)
        self.end_date = date.today()
        self._setup_ui()
        
    def _setup_ui(self):
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(3, weight=1)
        
        ctk.CTkLabel(self, text="De:", font=ctk.CTkFont(size=12)).grid(row=0, column=0, padx=5)
        self.start_entry = ctk.CTkEntry(self, width=110, height=32, font=ctk.CTkFont(size=12), placeholder_text="DD/MM/AAAA")
        self.start_entry.grid(row=0, column=1, padx=5)
        self.start_entry.insert(0, self.start_date.strftime("%d/%m/%Y"))
        self.start_entry.bind("<FocusOut>", lambda e: self._validate_dates())
        self.start_entry.bind("<Return>", lambda e: self._validate_dates())
        
        ctk.CTkLabel(self, text="Até:", font=ctk.CTkFont(size=12)).grid(row=0, column=2, padx=5)
        self.end_entry = ctk.CTkEntry(self, width=110, height=32, font=ctk.CTkFont(size=12), placeholder_text="DD/MM/AAAA")
        self.end_entry.grid(row=0, column=3, padx=5)
        self.end_entry.insert(0, self.end_date.strftime("%d/%m/%Y"))
        self.end_entry.bind("<FocusOut>", lambda e: self._validate_dates())
        self.end_entry.bind("<Return>", lambda e: self._validate_dates())
        
    def _validate_dates(self):
        try:
            d1, m1, y1 = map(int, self.start_entry.get().split("/"))
            d2, m2, y2 = map(int, self.end_entry.get().split("/"))
            self.start_date = date(y1, m1, d1)
            self.end_date = date(y2, m2, d2)
            if self.callback:
                self.callback(self.start_date, self.end_date)
        except (ValueError, AttributeError):
            pass
            
    def set_preset(self, preset: str):
        today = date.today()
        if preset == "today":
            self.start_date = self.end_date = today
        elif preset == "week":
            self.start_date = today - timedelta(days=today.weekday())
            self.end_date = today
        elif preset == "month":
            self.start_date = today.replace(day=1)
            self.end_date = today
        elif preset == "last_month":
            first = today.replace(day=1)
            self.end_date = first - timedelta(days=1)
            self.start_date = self.end_date.replace(day=1)
        elif preset == "year":
            self.start_date = today.replace(month=1, day=1)
            self.end_date = today
            
        self.start_entry.delete(0, "end")
        self.start_entry.insert(0, self.start_date.strftime("%d/%m/%Y"))
        self.end_entry.delete(0, "end")
        self.end_entry.insert(0, self.end_date.strftime("%d/%m/%Y"))
        
        if self.callback:
            self.callback(self.start_date, self.end_date)
            
    def get_dates(self) -> tuple[date, date]:
        return self.start_date, self.end_date


class ExportMixin:
    def export_to_excel(self, data: list, filename: str, sheet_name: str = "Dados") -> bool:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name[:31]
            
            if not data:
                wb.save(filename)
                return True
                
            if isinstance(data[0], dict):
                headers = list(data[0].keys())
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h.replace("_", " ").title())
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                    
                for row, item in enumerate(data, 2):
                    for col, h in enumerate(headers, 1):
                        val = item.get(h)
                        if hasattr(val, 'isoformat'):
                            val = val.isoformat()
                        ws.cell(row=row, column=col, value=val)
                        
            wb.save(filename)
            return True
        except Exception as e:
            print(f"Export error: {e}")
            return False
            
    def export_to_csv(self, data: list, filename: str) -> bool:
        try:
            import csv
            if not data or not isinstance(data[0], dict):
                return False
                
            with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=data[0].keys())
                writer.writeheader()
                for row in data:
                    writer.writerow({k: (v.isoformat() if hasattr(v, 'isoformat') else v) for k, v in row.items()})
            return True
        except Exception as e:
            print(f"CSV export error: {e}")
            return False


class StatusBadge(ctk.CTkLabel):
    COLORS = {
        "pending": "#f39c12",
        "received": "#27ae60",
        "cancelled": "#e74c3c",
        "open": "#3498db",
        "completed": "#27ae60",
        "returned": "#9b59b6",
        "paid": "#27ae60",
        "partial": "#f39c12",
        "refunded": "#e74c3c",
    }
    
    def __init__(self, parent, status: str, **kwargs):
        color = self.COLORS.get(status.lower(), "gray")
        super().__init__(parent, text=status.capitalize(), font=ctk.CTkFont(size=11, weight="bold"),
                        text_color=color, fg_color=color + "20", corner_radius=5, **kwargs)


class NumericEntry(ctk.CTkEntry):
    def __init__(self, parent, allow_float: bool = True, min_val: float = None, max_val: float = None, **kwargs):
        super().__init__(parent, **kwargs)
        self.allow_float = allow_float
        self.min_val = min_val
        self.max_val = max_val
        self.bind("<KeyRelease>", self._validate)
        self.bind("<FocusOut>", self._validate)
        
    def _validate(self, event=None):
        val = self.get()
        if not val:
            return True
            
        try:
            if self.allow_float:
                num = float(val)
            else:
                num = int(val)
                
            if self.min_val is not None and num < self.min_val:
                self.configure(border_color="#e74c3c")
                return False
            if self.max_val is not None and num > self.max_val:
                self.configure(border_color="#e74c3c")
                return False
                
            self.configure(border_color=("gray70", "gray30"))
            return True
        except ValueError:
            self.configure(border_color="#e74c3c")
            return False
            
    def get_value(self):
        val = self.get()
        if not val:
            return None
        try:
            return float(val) if self.allow_float else int(val)
        except ValueError:
            return None


class MoneyEntry(NumericEntry):
    def __init__(self, parent, **kwargs):
        super().__init__(parent, allow_float=True, min_val=0, **kwargs)
        self.bind("<FocusOut>", self._format)
        
    def _format(self, event=None):
        val = self.get_value()
        if val is not None:
            self.delete(0, "end")
            self.insert(0, f"{val:.2f}")


def center_window(window: ctk.CTk, width: int = None, height: int = None):
    window.update_idletasks()
    if width is None:
        width = window.winfo_width()
    if height is None:
        height = window.winfo_height()
    x = (window.winfo_screenwidth() // 2) - (width // 2)
    y = (window.winfo_screenheight() // 2) - (height // 2)
    window.geometry(f"{width}x{height}+{x}+{y}")