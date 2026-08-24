import customtkinter as ctk
from tkinter import messagebox, filedialog
from datetime import date, timedelta
from src.controllers.auth_controller import AuthController
from src.controllers.business_controllers import ReportController
from src.views.widgets import DateRangePicker, ExportMixin


class ReportsView(ctk.CTkFrame, ExportMixin):
    def __init__(self, parent, controller: ReportController, auth: AuthController):
        super().__init__(parent, fg_color="transparent")
        ExportMixin.__init__(self)
        self.controller = controller
        self.auth = auth
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.setup_ui()
        self.load_sales_report()
        
    def setup_ui(self):
        toolbar = ctk.CTkFrame(self, height=60, corner_radius=10)
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        
        self.tab_buttons = {}
        tabs = [
            ("sales", "📊 Vendas", self.show_sales_tab),
            ("inventory", "📦 Estoque", self.show_inventory_tab),
            ("financial", "💰 Financeiro", self.show_financial_tab),
            ("products", "📈 Produtos", self.show_products_tab),
        ]
        
        for i, (key, text, cmd) in enumerate(tabs):
            btn = ctk.CTkButton(toolbar, text=text, height=40,
                               font=ctk.CTkFont(size=13), corner_radius=8,
                               fg_color="transparent", text_color=("gray10", "gray90"),
                               hover_color=("gray80", "gray25"), command=cmd)
            btn.pack(side="left", padx=5, pady=10)
            self.tab_buttons[key] = btn
            
        ctk.CTkButton(toolbar, text="📤 Exportar", height=40, width=110,
                     command=self.export_current).pack(side="right", padx=15, pady=10)
        
        self.content_frame = ctk.CTkFrame(self, corner_radius=10)
        self.content_frame.grid(row=1, column=0, sticky="nsew")
        self.content_frame.grid_columnconfigure(0, weight=1)
        self.content_frame.grid_rowconfigure(0, weight=1)
        
        self.current_tab = "sales"
        self.current_data = []
        
    def set_active_tab(self, key: str):
        for k, btn in self.tab_buttons.items():
            if k == key:
                btn.configure(fg_color=("gray80", "gray25"), text_color=("blue", "lightblue"))
            else:
                btn.configure(fg_color="transparent", text_color=("gray10", "gray90"))
        self.current_tab = key
        
    def clear_content(self):
        for widget in self.content_frame.winfo_children():
            widget.destroy()
            
    def show_sales_tab(self):
        self.set_active_tab("sales")
        self.clear_content()
        
        top_frame = ctk.CTkFrame(self.content_frame, height=80, fg_color="transparent")
        top_frame.grid(row=0, column=0, sticky="ew", padx=20, pady=20)
        top_frame.grid_columnconfigure(2, weight=1)
        
        self.sales_date_picker = DateRangePicker(top_frame, callback=self.load_sales_report)
        self.sales_date_picker.grid(row=0, column=0, padx=10, sticky="w")
        
        ctk.CTkButton(top_frame, text="Hoje", width=80, height=35,
                     command=lambda: self.sales_date_picker.set_preset("today")).grid(row=0, column=1, padx=5)
        ctk.CTkButton(top_frame, text="Esta Semana", width=100, height=35,
                     command=lambda: self.sales_date_picker.set_preset("week")).grid(row=0, column=2, padx=5)
        ctk.CTkButton(top_frame, text="Este Mês", width=100, height=35,
                     command=lambda: self.sales_date_picker.set_preset("month")).grid(row=0, column=3, padx=5)
        
        self.sales_report_frame = ctk.CTkScrollableFrame(self.content_frame)
        self.sales_report_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=(0, 20))
        self.sales_report_frame.grid_columnconfigure(0, weight=1)
        
        self.load_sales_report()
        
    def load_sales_report(self, start_date=None, end_date=None):
        if start_date is None:
            start_date = self.sales_date_picker.start_date if hasattr(self, 'sales_date_picker') else date.today()
        if end_date is None:
            end_date = self.sales_date_picker.end_date if hasattr(self, 'sales_date_picker') else date.today()
            
        report = self.controller.get_sales_report(start_date, end_date)
        
        for widget in self.sales_report_frame.winfo_children():
            widget.destroy()
            
        summary = report.get("summary", {})
        cards_frame = ctk.CTkFrame(self.sales_report_frame, fg_color="transparent")
        cards_frame.grid(row=0, column=0, sticky="ew", pady=(0, 20))
        cards_frame.grid_columnconfigure((0,1,2,3), weight=1)
        
        cards = [
            ("Total Vendas", str(summary.get("total_sales", 0)), "#27ae60"),
            ("Receita Total", f"R$ {summary.get('total_revenue', 0):.2f}", "#2980b9"),
            ("Ticket Médio", f"R$ {summary.get('avg_ticket', 0):.2f}", "#8e44ad"),
            ("Descontos", f"R$ {summary.get('total_discount', 0):.2f}", "#e67e22"),
        ]
        
        for i, (title, value, color) in enumerate(cards):
            card = ctk.CTkFrame(cards_frame, corner_radius=10, border_width=2, border_color=color)
            card.grid(row=0, column=i, padx=10, sticky="nsew")
            ctk.CTkLabel(card, text=title, font=ctk.CTkFont(size=12), text_color="gray").pack(pady=(15, 5))
            ctk.CTkLabel(card, text=value, font=ctk.CTkFont(size=24, weight="bold"), text_color=color).pack(pady=(0, 15))
            
        pm_frame = ctk.CTkFrame(self.sales_report_frame, corner_radius=10)
        pm_frame.grid(row=1, column=0, sticky="ew", pady=(0, 20))
        pm_frame.grid_columnconfigure(0, weight=1)
        
        ctk.CTkLabel(pm_frame, text="💳 Formas de Pagamento", font=ctk.CTkFont(size=14, weight="bold")).grid(row=0, column=0, padx=20, pady=15, sticky="w")
        
        pm_data = report.get("payment_methods", [])
        for idx, pm in enumerate(pm_data):
            row = ctk.CTkFrame(pm_frame, fg_color=("gray90", "gray15") if idx % 2 == 0 else "transparent")
            row.grid(row=idx+1, column=0, sticky="ew", padx=10, pady=2)
            row.grid_columnconfigure(1, weight=1)
            ctk.CTkLabel(row, text=pm["payment_method"].replace("_", " ").title(), font=ctk.CTkFont(size=12)).grid(row=0, column=0, padx=15, pady=10, sticky="w")
            ctk.CTkLabel(row, text=f"{pm['count']} vendas", font=ctk.CTkFont(size=12), text_color="gray").grid(row=0, column=1, padx=15, pady=10)
            ctk.CTkLabel(row, text=f"R$ {pm['total']:.2f}", font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, column=2, padx=15, pady=10, sticky="e")
            
        tp_frame = ctk.CTkFrame(self.sales_report_frame, corner_radius=10)
        tp_frame.grid(row=2, column=0, sticky="nsew")
        tp_frame.grid_columnconfigure(0, weight=1)
        tp_frame.grid_rowconfigure(1, weight=1)
        
        ctk.CTkLabel(tp_frame, text="🏆 Top 10 Produtos Mais Vendidos", font=ctk.CTkFont(size=14, weight="bold")).grid(row=0, column=0, padx=20, pady=15, sticky="w")
        
        tp_data = report.get("top_products", [])
        if tp_data:
            cols = ["Produto", "Qtd Vendida", "Receita"]
            header = ctk.CTkFrame(tp_frame, fg_color=("gray85", "gray20"))
            header.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 5))
            for i, col in enumerate(cols):
                header.grid_columnconfigure(i, weight=1 if i == 0 else 0)
                ctk.CTkLabel(header, text=col, font=ctk.CTkFont(weight="bold")).grid(row=0, column=i, padx=15, pady=8)
                
            for idx, prod in enumerate(tp_data):
                row = ctk.CTkFrame(tp_frame, fg_color=("gray90", "gray15") if idx % 2 == 0 else "transparent")
                row.grid(row=idx+2, column=0, sticky="ew", padx=10, pady=1)
                row.grid_columnconfigure(0, weight=1)
                ctk.CTkLabel(row, text=prod["name"], font=ctk.CTkFont(size=12), anchor="w").grid(row=0, column=0, padx=15, pady=8, sticky="w")
                ctk.CTkLabel(row, text=str(prod["qty_sold"]), font=ctk.CTkFont(size=12)).grid(row=0, column=1, padx=15, pady=8)
                ctk.CTkLabel(row, text=f"R$ {prod['revenue']:.2f}", font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, column=2, padx=15, pady=8, sticky="e")
        else:
            ctk.CTkLabel(tp_frame, text="Nenhuma venda no período", text_color="gray").grid(row=1, column=0, pady=30)
            
        self.current_data = [("Resumo", summary), ("Formas Pagamento", pm_data), ("Top Produtos", tp_data)]
        
    def show_inventory_tab(self):
        self.set_active_tab("inventory")
        self.clear_content()
        
        ctk.CTkLabel(self.content_frame, text="📦 Relatório de Estoque", font=ctk.CTkFont(size=20, weight="bold")).pack(pady=20)
        
        inventory = self.controller.get_inventory_report()
        
        frame = ctk.CTkScrollableFrame(self.content_frame)
        frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        frame.grid_columnconfigure(0, weight=1)
        
        cols = ["SKU", "Produto", "Marca", "Categoria", "Estoque", "Mín", "Custo", "Venda", "Valor Estoque"]
        header = ctk.CTkFrame(frame, fg_color=("gray85", "gray20"))
        header.grid(row=0, column=0, sticky="ew", pady=(0, 5))
        for i, col in enumerate(cols):
            header.grid_columnconfigure(i, weight=1)
            ctk.CTkLabel(header, text=col, font=ctk.CTkFont(weight="bold", size=11)).grid(row=0, column=i, padx=8, pady=8)
            
        for idx, item in enumerate(inventory):
            row = ctk.CTkFrame(frame, fg_color=("gray90", "gray15") if idx % 2 == 0 else "transparent")
            row.grid(row=idx+1, column=0, sticky="ew", pady=1)
            row.grid_columnconfigure(0, weight=1)
            
            stock_val = item["current_stock"] * item["cost_price"]
            is_low = item["current_stock"] <= item["min_stock"]
            
            vals = [
                item["sku"], item["name"], item["brand"], item["category"],
                str(item["current_stock"]), str(item["min_stock"]),
                f"R$ {item['cost_price']:.2f}", f"R$ {item['sale_price']:.2f}",
                f"R$ {stock_val:.2f}"
            ]
            
            for i, val in enumerate(vals):
                color = "#e74c3c" if is_low and i == 4 else ("gray10", "gray90")
                ctk.CTkLabel(row, text=val, font=ctk.CTkFont(size=11), text_color=color, anchor="w").grid(row=0, column=i, padx=8, pady=6, sticky="w")
                
        self.current_data = [("Estoque", inventory)]
        
    def show_financial_tab(self):
        self.set_active_tab("financial")
        self.clear_content()
        
        top_frame = ctk.CTkFrame(self.content_frame, height=80, fg_color="transparent")
        top_frame.pack(fill="x", padx=20, pady=20)
        top_frame.grid_columnconfigure(2, weight=1)
        
        self.fin_date_picker = DateRangePicker(top_frame, callback=self.load_financial_report)
        self.fin_date_picker.grid(row=0, column=0, padx=10, sticky="w")
        
        ctk.CTkButton(top_frame, text="Este Mês", width=100, height=35,
                     command=lambda: self.fin_date_picker.set_preset("month")).grid(row=0, column=1, padx=5)
        ctk.CTkButton(top_frame, text="Mês Passado", width=100, height=35,
                     command=lambda: self.fin_date_picker.set_preset("last_month")).grid(row=0, column=2, padx=5)
        
        self.fin_frame = ctk.CTkScrollableFrame(self.content_frame)
        self.fin_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        self.load_financial_report()
        
    def load_financial_report(self, start_date=None, end_date=None):
        if start_date is None:
            start_date = self.fin_date_picker.start_date if hasattr(self, 'fin_date_picker') else date.today().replace(day=1)
        if end_date is None:
            end_date = self.fin_date_picker.end_date if hasattr(self, 'fin_date_picker') else date.today()
            
        fin = self.controller.get_financial_summary(start_date, end_date)
        
        for widget in self.fin_frame.winfo_children():
            widget.destroy()
            
        cards_frame = ctk.CTkFrame(self.fin_frame, fg_color="transparent")
        cards_frame.pack(fill="x", pady=(0, 20))
        cards_frame.grid_columnconfigure((0,1,2), weight=1)
        
        cards = [
            ("💰 Vendas Totais", f"R$ {fin['total_sales']:.2f}", "#27ae60"),
            ("📦 Compras Totais", f"R$ {fin['total_purchases']:.2f}", "#e74c3c"),
            ("📊 Lucro Bruto", f"R$ {fin['gross_profit']:.2f}", "#2980b9" if fin['gross_profit'] >= 0 else "#e74c3c"),
        ]
        
        for i, (title, value, color) in enumerate(cards):
            card = ctk.CTkFrame(cards_frame, corner_radius=15, border_width=2, border_color=color)
            card.grid(row=0, column=i, padx=10, sticky="nsew")
            ctk.CTkLabel(card, text=title, font=ctk.CTkFont(size=13), text_color="gray").pack(pady=(20, 5))
            ctk.CTkLabel(card, text=value, font=ctk.CTkFont(size=28, weight="bold"), text_color=color).pack(pady=(0, 20))
            
        margin = (fin['gross_profit'] / fin['total_sales'] * 100) if fin['total_sales'] > 0 else 0
        margin_card = ctk.CTkFrame(self.fin_frame, corner_radius=15, border_width=2, border_color="#f39c12")
        margin_card.pack(fill="x", pady=10)
        ctk.CTkLabel(margin_card, text="📈 Margem de Lucro Bruto", font=ctk.CTkFont(size=14), text_color="gray").pack(pady=(15, 5))
        ctk.CTkLabel(margin_card, text=f"{margin:.1f}%", font=ctk.CTkFont(size=36, weight="bold"), text_color="#f39c12").pack(pady=(0, 15))
        
        self.current_data = [("Financeiro", fin)]
        
    def show_products_tab(self):
        self.set_active_tab("products")
        self.clear_content()
        
        ctk.CTkLabel(self.content_frame, text="📈 Análise de Produtos", font=ctk.CTkFont(size=20, weight="bold")).pack(pady=20)
        
        frame = ctk.CTkFrame(self.content_frame, corner_radius=10)
        frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        ctk.CTkLabel(frame, text="📊 Gráficos e análises avançadas\n(Implementar com matplotlib/chart.js)",
                    font=ctk.CTkFont(size=14), text_color="gray").pack(pady=50)
        
    def export_current(self):
        if not self.current_data:
            messagebox.showwarning("Aviso", "Nenhum dado para exportar")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")],
            initialname=f"relatorio_{self.current_tab}_{date.today().strftime('%Y%m%d')}.xlsx"
        )
        if not file_path:
            return
            
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            
            for sheet_name, data in self.current_data:
                if sheet_name == "Resumo" and isinstance(data, dict):
                    ws = wb.active
                    ws.title = "Resumo"
                    for row, (k, v) in enumerate(data.items(), 1):
                        ws.cell(row=row, column=1, value=k.replace("_", " ").title())
                        ws.cell(row=row, column=2, value=v)
                elif isinstance(data, list) and data and isinstance(data[0], dict):
                    ws = wb.create_sheet(title=sheet_name[:31])
                    headers = list(data[0].keys())
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h.replace("_", " ").title())
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    for row, item in enumerate(data, 2):
                        for col, h in enumerate(headers, 1):
                            ws.cell(row=row, column=col, value=item.get(h))
                            
            if len(wb.sheetnames) > 1:
                wb.remove(wb[wb.sheetnames[0]])
                
            wb.save(file_path)
            messagebox.showinfo("Sucesso", f"Exportado para {file_path}")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao exportar: {e}")